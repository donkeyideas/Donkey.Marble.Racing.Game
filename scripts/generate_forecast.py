"""
Donkey Marble Racing — 60-Month Financial Forecast Generator
Generates a comprehensive Excel workbook with pricing, revenue forecasts,
app store fee analysis, and business KPIs.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference, PieChart
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.label import DataLabelList
from copy import copy
import math

wb = openpyxl.Workbook()

# ── Color Palette ──
DARK_BG = "0A1A3A"
GOLD = "FFC220"
BLUE = "6EC1FF"
GREEN = "2ECC71"
RED = "E74C3C"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
MED_GRAY = "D9D9D9"
HEADER_BG = "1A2E5A"
ROW_ALT = "EBF5FB"
SECTION_BG = "2C3E6B"

# ── Reusable styles ──
header_font = Font(name="Calibri", bold=True, color=WHITE, size=12)
title_font = Font(name="Calibri", bold=True, color=DARK_BG, size=14)
section_font = Font(name="Calibri", bold=True, color=GOLD, size=12)
money_fmt = '#,##0.00'
int_fmt = '#,##0'
pct_fmt = '0.0%'
usd_fmt = '$#,##0.00'
usd_whole_fmt = '$#,##0'

header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
gold_fill = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
alt_fill = PatternFill(start_color=ROW_ALT, end_color=ROW_ALT, fill_type="solid")
section_fill = PatternFill(start_color=SECTION_BG, end_color=SECTION_BG, fill_type="solid")
green_fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
red_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
light_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color=MED_GRAY),
    right=Side(style='thin', color=MED_GRAY),
    top=Side(style='thin', color=MED_GRAY),
    bottom=Side(style='thin', color=MED_GRAY),
)


def style_header_row(ws, row, max_col, fill=None, font=None):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill or header_fill
        cell.font = font or header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border


def style_data_cell(cell, fmt=None, alt=False):
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        cell.number_format = fmt
    if alt:
        cell.fill = alt_fill


def write_row(ws, row, data, start_col=1, fmt=None, alt=False, bold=False):
    for i, val in enumerate(data):
        cell = ws.cell(row=row, column=start_col + i, value=val)
        style_data_cell(cell, fmt=fmt if isinstance(fmt, str) else (fmt[i] if fmt else None), alt=alt)
        if bold:
            cell.font = Font(bold=True)


# ═══════════════════════════════════════════════════════════════
# SHEET 1: DASHBOARD
# ═══════════════════════════════════════════════════════════════
ws_dash = wb.active
ws_dash.title = "Dashboard"
ws_dash.sheet_properties.tabColor = GOLD

# Title
ws_dash.merge_cells("A1:H1")
cell = ws_dash["A1"]
cell.value = "DONKEY MARBLE RACING — FINANCIAL FORECAST DASHBOARD"
cell.font = Font(name="Calibri", bold=True, color=DARK_BG, size=18)
cell.alignment = Alignment(horizontal="center", vertical="center")
cell.fill = gold_fill

ws_dash.merge_cells("A2:H2")
cell = ws_dash["A2"]
cell.value = "60-Month Business Plan & Revenue Projections"
cell.font = Font(name="Calibri", color="666666", size=12)
cell.alignment = Alignment(horizontal="center")

# Key Assumptions Box
r = 4
ws_dash.merge_cells(f"A{r}:D{r}")
ws_dash[f"A{r}"].value = "KEY ASSUMPTIONS"
ws_dash[f"A{r}"].font = section_font
ws_dash[f"A{r}"].fill = section_fill
style_header_row(ws_dash, r, 4, fill=section_fill, font=section_font)

assumptions = [
    ("Launch Platform", "iOS + Android (simultaneous)"),
    ("Apple App Store Fee", "30% (15% under Small Business Program < $1M/yr)"),
    ("Google Play Fee", "15% (first $1M/yr), then 30%"),
    ("Avg. Conversion Rate (Free→Paid)", "3-5% (industry avg for casual games)"),
    ("Avg. Revenue Per Paying User (ARPPU)", "$8-15/month"),
    ("Monthly Churn Rate", "8-12% for paying users"),
    ("User Acquisition Cost (est.)", "$1.50-3.00 per install"),
    ("Organic Install Rate", "40% of total installs"),
    ("Season Length", "30 days"),
    ("Daily Active User Rate", "25-35% of total installs"),
]
for i, (label, value) in enumerate(assumptions):
    row = r + 1 + i
    ws_dash.cell(row=row, column=1, value=label).font = Font(bold=True)
    ws_dash.cell(row=row, column=1).border = thin_border
    ws_dash.merge_cells(f"B{row}:D{row}")
    ws_dash.cell(row=row, column=2, value=value).border = thin_border
    if i % 2 == 0:
        for c in range(1, 5):
            ws_dash.cell(row=row, column=c).fill = alt_fill

# Revenue Streams Box
r2 = 4
ws_dash.merge_cells(f"F{r2}:H{r2}")
ws_dash[f"F{r2}"].value = "REVENUE STREAMS"
ws_dash[f"F{r2}"].font = section_font
ws_dash[f"F{r2}"].fill = section_fill
style_header_row(ws_dash, r2, 8, fill=section_fill, font=section_font)

streams = [
    ("Coin Packs (IAP)", "$0.99 - $24.99", "Primary"),
    ("Season Pass — Premium", "$9.99/season", "Recurring"),
    ("Season Pass — Plus", "$24.99/season", "Recurring"),
    ("Future: Ad Revenue (rewarded)", "eCPM $10-30", "Supplemental"),
    ("Future: Cosmetic Store", "$0.99 - $4.99", "Supplemental"),
    ("Future: Tournament Entry", "Coin-gated", "Engagement"),
]
for i, (stream, price, category) in enumerate(streams):
    row = r2 + 1 + i
    ws_dash.cell(row=row, column=6, value=stream).font = Font(bold=True)
    ws_dash.cell(row=row, column=6).border = thin_border
    ws_dash.cell(row=row, column=7, value=price).border = thin_border
    ws_dash.cell(row=row, column=8, value=category).border = thin_border
    if i % 2 == 0:
        for c in range(6, 9):
            ws_dash.cell(row=row, column=c).fill = alt_fill

# Column widths
for col in range(1, 9):
    ws_dash.column_dimensions[get_column_letter(col)].width = 22


# ═══════════════════════════════════════════════════════════════
# SHEET 2: PRICING MATRIX (After App Store Fees)
# ═══════════════════════════════════════════════════════════════
ws_price = wb.create_sheet("Pricing Matrix")
ws_price.sheet_properties.tabColor = GREEN

# Title
ws_price.merge_cells("A1:J1")
ws_price["A1"].value = "PRICING MATRIX — REVENUE AFTER APP STORE FEES"
ws_price["A1"].font = Font(name="Calibri", bold=True, color=WHITE, size=16)
ws_price["A1"].fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
ws_price["A1"].alignment = Alignment(horizontal="center")

# Section A: Current Products
r = 3
ws_price.merge_cells(f"A{r}:J{r}")
ws_price[f"A{r}"].value = "A. CURRENT IN-APP PURCHASES"
style_header_row(ws_price, r, 10, fill=section_fill, font=section_font)

headers = [
    "Product", "Retail Price", "Coins Given", "Coins/Dollar",
    "Apple 30%\nYou Receive", "Apple 15%\n(Small Biz)\nYou Receive",
    "Google 15%\nYou Receive", "Google 30%\nYou Receive",
    "Best Case\n(15% fee)", "Worst Case\n(30% fee)"
]
r += 1
for i, h in enumerate(headers):
    cell = ws_price.cell(row=r, column=i + 1, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

products = [
    ("Starter Pack", 0.99, 1000),
    ("Popular Pack", 4.99, 6000),
    ("Big Spender", 9.99, 15000),
    ("Whale Pack", 24.99, 40000),
    ("Season Pass — Premium", 9.99, None),
    ("Season Pass — Plus", 24.99, None),
]

for i, (name, price, coins) in enumerate(products):
    row = r + 1 + i
    alt = i % 2 == 0

    coins_per_dollar = round(coins / price, 0) if coins else "N/A"
    apple_30 = round(price * 0.70, 2)
    apple_15 = round(price * 0.85, 2)
    google_15 = round(price * 0.85, 2)
    google_30 = round(price * 0.70, 2)
    best = round(price * 0.85, 2)
    worst = round(price * 0.70, 2)

    data = [name, price, coins if coins else "—", coins_per_dollar,
            apple_30, apple_15, google_15, google_30, best, worst]
    fmts = [None, usd_fmt, int_fmt if coins else None, int_fmt if coins else None,
            usd_fmt, usd_fmt, usd_fmt, usd_fmt, usd_fmt, usd_fmt]
    write_row(ws_price, row, data, fmt=fmts, alt=alt)
    ws_price.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")

# Section B: Fee Comparison Table
r_fee = r + len(products) + 3
ws_price.merge_cells(f"A{r_fee}:J{r_fee}")
ws_price[f"A{r_fee}"].value = "B. APP STORE FEE COMPARISON"
style_header_row(ws_price, r_fee, 10, fill=section_fill, font=section_font)

r_fee += 1
fee_headers = ["Scenario", "Fee Rate", "On $0.99", "On $4.99", "On $9.99", "On $24.99",
               "On $100 Revenue", "On $1,000 Revenue", "On $10,000 Revenue", "Annual Impact"]
for i, h in enumerate(fee_headers):
    cell = ws_price.cell(row=r_fee, column=i + 1, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

fee_scenarios = [
    ("Apple Standard", 0.30),
    ("Apple Small Business (<$1M)", 0.15),
    ("Google Play (first $1M)", 0.15),
    ("Google Play (over $1M)", 0.30),
]

test_prices = [0.99, 4.99, 9.99, 24.99]
test_revenues = [100, 1000, 10000]

for i, (scenario, rate) in enumerate(fee_scenarios):
    row = r_fee + 1 + i
    alt = i % 2 == 0
    data = [scenario, rate]
    fmts_list = [None, pct_fmt]
    for p in test_prices:
        data.append(round(p * rate, 2))
        fmts_list.append(usd_fmt)
    for rev in test_revenues:
        data.append(round(rev * rate, 2))
        fmts_list.append(usd_fmt)
    # Annual impact on hypothetical $500k
    data.append(round(500000 * rate, 0))
    fmts_list.append(usd_whole_fmt)
    write_row(ws_price, row, data, fmt=fmts_list, alt=alt)

# Section C: Break-even analysis
r_be = r_fee + len(fee_scenarios) + 3
ws_price.merge_cells(f"A{r_be}:F{r_be}")
ws_price[f"A{r_be}"].value = "C. UNITS NEEDED TO REACH REVENUE TARGETS (After 15% Fee)"
style_header_row(ws_price, r_be, 6, fill=section_fill, font=section_font)

r_be += 1
be_headers = ["Product", "Net Per Unit", "Units for $1K/mo", "Units for $5K/mo", "Units for $10K/mo", "Units for $50K/mo"]
for i, h in enumerate(be_headers):
    cell = ws_price.cell(row=r_be, column=i + 1, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

targets = [1000, 5000, 10000, 50000]
for i, (name, price, coins) in enumerate(products):
    row = r_be + 1 + i
    net = round(price * 0.85, 2)
    data = [name, net]
    fmts_list = [None, usd_fmt]
    for t in targets:
        units = math.ceil(t / net)
        data.append(units)
        fmts_list.append(int_fmt)
    write_row(ws_price, row, data, fmt=fmts_list, alt=i % 2 == 0)

# Column widths
for col in range(1, 11):
    ws_price.column_dimensions[get_column_letter(col)].width = 18
ws_price.column_dimensions["A"].width = 28


# ═══════════════════════════════════════════════════════════════
# SHEET 3: 60-MONTH FORECAST
# ═══════════════════════════════════════════════════════════════
ws_forecast = wb.create_sheet("60-Month Forecast")
ws_forecast.sheet_properties.tabColor = BLUE

ws_forecast.merge_cells("A1:G1")
ws_forecast["A1"].value = "60-MONTH REVENUE & USER FORECAST"
ws_forecast["A1"].font = Font(name="Calibri", bold=True, color=WHITE, size=16)
ws_forecast["A1"].fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
ws_forecast["A1"].alignment = Alignment(horizontal="center")

# Input parameters (editable by user)
r = 3
ws_forecast.merge_cells(f"A{r}:C{r}")
ws_forecast[f"A{r}"].value = "EDITABLE INPUTS (Change yellow cells to update forecast)"
ws_forecast[f"A{r}"].font = section_font
ws_forecast[f"A{r}"].fill = section_fill
style_header_row(ws_forecast, r, 3, fill=section_fill, font=section_font)

inputs = [
    ("Launch Month Installs", 2000, "B"),
    ("Monthly Install Growth Rate", 0.12, "B"),
    ("Install Growth Decay (per month)", 0.003, "B"),
    ("Free-to-Paid Conversion Rate", 0.04, "B"),
    ("Avg Revenue Per Paying User (ARPPU)", 7.50, "B"),
    ("Monthly Paying User Churn Rate", 0.10, "B"),
    ("Monthly Free User Churn Rate", 0.15, "B"),
    ("ARPPU Annual Growth Rate", 0.05, "B"),
    ("App Store Fee Rate", 0.15, "B"),
    ("Monthly Server/Infra Cost", 200, "B"),
    ("Monthly Marketing Spend", 500, "B"),
    ("Cost Per Install (paid)", 2.00, "B"),
    ("Organic Install %", 0.40, "B"),
]

input_cells = {}
for i, (label, default, col) in enumerate(inputs):
    row = r + 1 + i
    cell_a = ws_forecast.cell(row=row, column=1, value=label)
    cell_a.font = Font(bold=True)
    cell_a.border = thin_border
    cell_a.alignment = Alignment(horizontal="left")
    if i % 2 == 0:
        cell_a.fill = alt_fill

    cell_b = ws_forecast.cell(row=row, column=2, value=default)
    cell_b.fill = gold_fill
    cell_b.font = Font(bold=True)
    cell_b.border = thin_border
    cell_b.alignment = Alignment(horizontal="center")

    if isinstance(default, float) and default < 1:
        cell_b.number_format = pct_fmt
    elif isinstance(default, float):
        cell_b.number_format = usd_fmt
    else:
        cell_b.number_format = int_fmt

    input_cells[label] = f"B{row}"

# Map input cells for formulas
launch_installs = input_cells["Launch Month Installs"]
growth_rate = input_cells["Monthly Install Growth Rate"]
growth_decay = input_cells["Install Growth Decay (per month)"]
conversion_rate = input_cells["Free-to-Paid Conversion Rate"]
arppu = input_cells["Avg Revenue Per Paying User (ARPPU)"]
churn_paid = input_cells["Monthly Paying User Churn Rate"]
churn_free = input_cells["Monthly Free User Churn Rate"]
arppu_growth = input_cells["ARPPU Annual Growth Rate"]
fee_rate = input_cells["App Store Fee Rate"]
server_cost = input_cells["Monthly Server/Infra Cost"]
marketing_spend = input_cells["Monthly Marketing Spend"]
cpi = input_cells["Cost Per Install (paid)"]
organic_pct = input_cells["Organic Install %"]

# Forecast table
fr = r + len(inputs) + 2
forecast_headers = [
    "Month", "New Installs", "Cumulative Installs",
    "Total Free Users", "Total Paying Users", "Total Active Users",
    "New Conversions", "Gross Revenue", "App Store Fees",
    "Net Revenue", "Server Costs", "Marketing Spend",
    "UA Cost (Paid Installs)", "Total Costs",
    "Net Profit/Loss", "Cumulative Profit/Loss",
    "ARPU (All Users)", "LTV Est. (Paying)",
    "Paying User %"
]

ws_forecast.merge_cells(f"A{fr}:S{fr}")
ws_forecast[f"A{fr}"].value = "MONTHLY FORECAST"
style_header_row(ws_forecast, fr, 19, fill=section_fill, font=section_font)

fr += 1
for i, h in enumerate(forecast_headers):
    cell = ws_forecast.cell(row=fr, column=i + 1, value=h)
    cell.fill = header_fill
    cell.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

ws_forecast.row_dimensions[fr].height = 45

# Generate 60 months with Excel formulas
for m in range(1, 61):
    row = fr + m
    alt = m % 2 == 0

    # A: Month number
    ws_forecast.cell(row=row, column=1, value=m)

    # B: New Installs = launch * (1 + growth - decay*month)^(month-1)
    if m == 1:
        ws_forecast.cell(row=row, column=2).value = f"={launch_installs}"
    else:
        ws_forecast.cell(row=row, column=2).value = \
            f"=ROUND({launch_installs}*(1+{growth_rate}-{growth_decay}*{m})^({m}-1),0)"

    # C: Cumulative Installs
    if m == 1:
        ws_forecast.cell(row=row, column=3).value = f"=B{row}"
    else:
        ws_forecast.cell(row=row, column=3).value = f"=C{row-1}+B{row}"

    # D: Total Free Users = prev_free*(1-churn_free) + new_installs*(1-conversion)
    if m == 1:
        ws_forecast.cell(row=row, column=4).value = f"=ROUND(B{row}*(1-{conversion_rate}),0)"
    else:
        ws_forecast.cell(row=row, column=4).value = \
            f"=ROUND(D{row-1}*(1-{churn_free})+B{row}*(1-{conversion_rate}),0)"

    # E: Total Paying Users = prev_paying*(1-churn_paid) + new_conversions
    if m == 1:
        ws_forecast.cell(row=row, column=5).value = f"=ROUND(B{row}*{conversion_rate},0)"
    else:
        ws_forecast.cell(row=row, column=5).value = \
            f"=ROUND(E{row-1}*(1-{churn_paid})+G{row},0)"

    # F: Total Active Users
    ws_forecast.cell(row=row, column=6).value = f"=D{row}+E{row}"

    # G: New Conversions this month
    if m == 1:
        ws_forecast.cell(row=row, column=7).value = f"=ROUND(B{row}*{conversion_rate},0)"
    else:
        ws_forecast.cell(row=row, column=7).value = \
            f"=ROUND(B{row}*{conversion_rate}+D{row-1}*0.005,0)"

    # H: Gross Revenue = paying_users * ARPPU * (1 + annual_growth)^(year)
    year_mult = f"(1+{arppu_growth})^(INT(({m}-1)/12))"
    ws_forecast.cell(row=row, column=8).value = f"=ROUND(E{row}*{arppu}*{year_mult},2)"

    # I: App Store Fees
    ws_forecast.cell(row=row, column=9).value = f"=ROUND(H{row}*{fee_rate},2)"

    # J: Net Revenue
    ws_forecast.cell(row=row, column=10).value = f"=H{row}-I{row}"

    # K: Server Costs (scales with users)
    ws_forecast.cell(row=row, column=11).value = \
        f"=ROUND({server_cost}*(1+F{row}/10000*0.5),2)"

    # L: Marketing Spend
    ws_forecast.cell(row=row, column=12).value = f"={marketing_spend}"

    # M: UA Cost (paid installs only)
    ws_forecast.cell(row=row, column=13).value = \
        f"=ROUND(B{row}*(1-{organic_pct})*{cpi},2)"

    # N: Total Costs
    ws_forecast.cell(row=row, column=14).value = f"=K{row}+L{row}+M{row}"

    # O: Net Profit/Loss
    ws_forecast.cell(row=row, column=15).value = f"=J{row}-N{row}"

    # P: Cumulative P/L
    if m == 1:
        ws_forecast.cell(row=row, column=16).value = f"=O{row}"
    else:
        ws_forecast.cell(row=row, column=16).value = f"=P{row-1}+O{row}"

    # Q: ARPU (all users)
    ws_forecast.cell(row=row, column=17).value = f"=IF(F{row}>0,ROUND(J{row}/F{row},2),0)"

    # R: LTV Estimate (paying user)
    ws_forecast.cell(row=row, column=18).value = \
        f"=IF({churn_paid}>0,ROUND({arppu}*(1-{fee_rate})/{churn_paid}*{year_mult},2),0)"

    # S: Paying User %
    ws_forecast.cell(row=row, column=19).value = f"=IF(F{row}>0,E{row}/F{row},0)"

    # Style all cells
    for col in range(1, 20):
        cell = ws_forecast.cell(row=row, column=col)
        cell.border = thin_border
        if alt:
            cell.fill = alt_fill
        cell.alignment = Alignment(horizontal="center")

    # Number formats
    ws_forecast.cell(row=row, column=1).number_format = '0'
    ws_forecast.cell(row=row, column=2).number_format = int_fmt
    ws_forecast.cell(row=row, column=3).number_format = int_fmt
    ws_forecast.cell(row=row, column=4).number_format = int_fmt
    ws_forecast.cell(row=row, column=5).number_format = int_fmt
    ws_forecast.cell(row=row, column=6).number_format = int_fmt
    ws_forecast.cell(row=row, column=7).number_format = int_fmt
    for col in range(8, 16):
        ws_forecast.cell(row=row, column=col).number_format = usd_fmt
    ws_forecast.cell(row=row, column=16).number_format = usd_fmt
    ws_forecast.cell(row=row, column=17).number_format = usd_fmt
    ws_forecast.cell(row=row, column=18).number_format = usd_fmt
    ws_forecast.cell(row=row, column=19).number_format = pct_fmt

# Summary row
summary_row = fr + 61
ws_forecast.cell(row=summary_row, column=1, value="TOTALS").font = Font(bold=True, size=12)
for col_idx, col_letter in [(3, "C"), (8, "H"), (9, "I"), (10, "J"),
                             (11, "K"), (12, "L"), (13, "M"), (14, "N"), (15, "O")]:
    cell = ws_forecast.cell(row=summary_row, column=col_idx)
    cell.value = f"=SUM({col_letter}{fr+1}:{col_letter}{fr+60})"
    cell.number_format = usd_fmt if col_idx >= 8 else int_fmt
    cell.font = Font(bold=True, size=12)
    cell.fill = gold_fill
    cell.border = thin_border

# Column widths
col_widths = [8, 14, 16, 15, 15, 15, 14, 15, 14, 14, 13, 14, 16, 13, 15, 17, 14, 16, 13]
for i, w in enumerate(col_widths):
    ws_forecast.column_dimensions[get_column_letter(i + 1)].width = w

# Freeze panes
ws_forecast.freeze_panes = f"A{fr+1}"


# ═══════════════════════════════════════════════════════════════
# SHEET 4: REVENUE BY STREAM
# ═══════════════════════════════════════════════════════════════
ws_rev = wb.create_sheet("Revenue Streams")
ws_rev.sheet_properties.tabColor = "E74C3C"

ws_rev.merge_cells("A1:H1")
ws_rev["A1"].value = "REVENUE BREAKDOWN BY STREAM — QUARTERLY PROJECTIONS"
ws_rev["A1"].font = Font(name="Calibri", bold=True, color=WHITE, size=16)
ws_rev["A1"].fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
ws_rev["A1"].alignment = Alignment(horizontal="center")

# Revenue mix assumptions
r = 3
ws_rev.merge_cells(f"A{r}:H{r}")
ws_rev[f"A{r}"].value = "REVENUE MIX ASSUMPTIONS (% of Gross Revenue)"
style_header_row(ws_rev, r, 8, fill=section_fill, font=section_font)

r += 1
mix_headers = ["Revenue Stream", "Year 1 Mix", "Year 2 Mix", "Year 3 Mix", "Year 4 Mix", "Year 5 Mix",
               "Avg. Transaction", "Notes"]
for i, h in enumerate(mix_headers):
    cell = ws_rev.cell(row=r, column=i + 1, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

mix_data = [
    ("Coin Pack — Starter ($0.99)", 0.15, 0.10, 0.08, 0.06, 0.05, 0.99, "Entry-level, high volume"),
    ("Coin Pack — Popular ($4.99)", 0.25, 0.22, 0.20, 0.18, 0.15, 4.99, "Best value perception"),
    ("Coin Pack — Big Spender ($9.99)", 0.15, 0.18, 0.18, 0.17, 0.15, 9.99, "Mid-tier spenders"),
    ("Coin Pack — Whale ($24.99)", 0.05, 0.08, 0.10, 0.12, 0.15, 24.99, "Whale segment grows"),
    ("Season Pass — Premium ($9.99)", 0.25, 0.22, 0.20, 0.20, 0.18, 9.99, "Core recurring"),
    ("Season Pass — Plus ($24.99)", 0.10, 0.12, 0.14, 0.15, 0.17, 24.99, "Premium recurring"),
    ("Cosmetics/Skins (Future)", 0.00, 0.05, 0.07, 0.08, 0.10, 2.99, "Phase C+"),
    ("Rewarded Ads (Future)", 0.05, 0.03, 0.03, 0.04, 0.05, 0.02, "Per impression eCPM"),
]

for i, row_data in enumerate(mix_data):
    row = r + 1 + i
    name, *years, avg_txn, notes = row_data
    data = [name] + list(years) + [avg_txn, notes]
    fmts = [None, pct_fmt, pct_fmt, pct_fmt, pct_fmt, pct_fmt, usd_fmt, None]
    write_row(ws_rev, row, data, fmt=fmts, alt=i % 2 == 0)
    ws_rev.cell(row=row, column=1).alignment = Alignment(horizontal="left")

# Totals row
total_row = r + 1 + len(mix_data)
ws_rev.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
for col in range(2, 7):
    cell = ws_rev.cell(row=total_row, column=col)
    cell.value = f"=SUM({get_column_letter(col)}{r+1}:{get_column_letter(col)}{total_row-1})"
    cell.number_format = pct_fmt
    cell.font = Font(bold=True)
    cell.fill = gold_fill
    cell.border = thin_border

# Quarterly breakdown
qr = total_row + 3
ws_rev.merge_cells(f"A{qr}:H{qr}")
ws_rev[f"A{qr}"].value = "QUARTERLY REVENUE BY STREAM (Estimated)"
style_header_row(ws_rev, qr, 8, fill=section_fill, font=section_font)

qr += 1
q_headers = ["Quarter", "Coin Packs", "Season Pass Premium", "Season Pass Plus",
             "Cosmetics", "Ads", "Total Gross", "Total Net (after 15% fee)"]
for i, h in enumerate(q_headers):
    cell = ws_rev.cell(row=qr, column=i + 1, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

# Estimated quarterly revenue based on growth curve
quarterly_gross = [
    ("Q1 Y1", 450, 375, 150, 0, 75),
    ("Q2 Y1", 900, 660, 300, 0, 90),
    ("Q3 Y1", 1800, 1100, 540, 0, 110),
    ("Q4 Y1", 3200, 1760, 880, 0, 160),
    ("Q1 Y2", 5000, 2500, 1350, 350, 200),
    ("Q2 Y2", 7500, 3300, 1800, 600, 225),
    ("Q3 Y2", 10500, 4200, 2400, 900, 315),
    ("Q4 Y2", 14000, 5040, 3080, 1260, 420),
    ("Q1 Y3", 17500, 5600, 3850, 1750, 525),
    ("Q2 Y3", 21000, 6300, 4410, 2100, 630),
    ("Q3 Y3", 24500, 6860, 4900, 2450, 735),
    ("Q4 Y3", 28000, 7280, 5320, 2800, 840),
    ("Q1 Y4", 31000, 7750, 5580, 3100, 930),
    ("Q2 Y4", 33500, 8040, 5695, 3350, 1005),
    ("Q3 Y4", 35500, 8165, 5680, 3550, 1065),
    ("Q4 Y4", 37000, 8140, 5550, 3700, 1110),
    ("Q1 Y5", 38000, 7980, 5320, 3800, 1140),
    ("Q2 Y5", 38500, 7700, 5005, 3850, 1155),
    ("Q3 Y5", 38500, 7315, 4620, 3850, 1155),
    ("Q4 Y5", 38000, 6840, 4180, 3800, 1140),
]

for i, (q, coins, prem, plus, cosm, ads) in enumerate(quarterly_gross):
    row = qr + 1 + i
    total = coins + prem + plus + cosm + ads
    net = round(total * 0.85, 2)
    data = [q, coins, prem, plus, cosm, ads, total, net]
    fmts = [None, usd_whole_fmt, usd_whole_fmt, usd_whole_fmt, usd_whole_fmt,
            usd_whole_fmt, usd_whole_fmt, usd_whole_fmt]
    write_row(ws_rev, row, data, fmt=fmts, alt=i % 2 == 0)
    ws_rev.cell(row=row, column=1).font = Font(bold=True)

# Grand total
gt_row = qr + 1 + len(quarterly_gross)
ws_rev.cell(row=gt_row, column=1, value="5-YEAR TOTAL").font = Font(bold=True, size=12)
for col in range(2, 9):
    cell = ws_rev.cell(row=gt_row, column=col)
    cell.value = f"=SUM({get_column_letter(col)}{qr+1}:{get_column_letter(col)}{gt_row-1})"
    cell.number_format = usd_whole_fmt
    cell.font = Font(bold=True, size=12)
    cell.fill = gold_fill
    cell.border = thin_border

# Column widths
ws_rev.column_dimensions["A"].width = 30
for col in range(2, 9):
    ws_rev.column_dimensions[get_column_letter(col)].width = 20


# ═══════════════════════════════════════════════════════════════
# SHEET 5: UNIT ECONOMICS & KPIs
# ═══════════════════════════════════════════════════════════════
ws_kpi = wb.create_sheet("KPIs & Unit Economics")
ws_kpi.sheet_properties.tabColor = "9B59B6"

ws_kpi.merge_cells("A1:F1")
ws_kpi["A1"].value = "KPIs & UNIT ECONOMICS"
ws_kpi["A1"].font = Font(name="Calibri", bold=True, color=WHITE, size=16)
ws_kpi["A1"].fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
ws_kpi["A1"].alignment = Alignment(horizontal="center")

# Key Metrics
r = 3
ws_kpi.merge_cells(f"A{r}:F{r}")
ws_kpi[f"A{r}"].value = "KEY BUSINESS METRICS"
style_header_row(ws_kpi, r, 6, fill=section_fill, font=section_font)

r += 1
kpi_headers = ["Metric", "Formula", "Year 1", "Year 2", "Year 3", "Benchmark"]
for i, h in enumerate(kpi_headers):
    cell = ws_kpi.cell(row=r, column=i + 1, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

kpis = [
    ("Customer Acquisition Cost (CAC)", "Marketing Spend / New Users", "$2.00", "$1.80", "$1.50", "$1.50-3.00"),
    ("Lifetime Value (LTV) — Paying", "ARPPU / Churn Rate", "$63.75", "$66.94", "$70.28", "$30-80"),
    ("LTV:CAC Ratio", "LTV / CAC", "31.9x", "37.2x", "46.9x", "> 3:1 healthy"),
    ("Monthly ARPU (All Users)", "Revenue / Active Users", "$0.28", "$0.45", "$0.62", "$0.20-1.00"),
    ("Monthly ARPPU (Paying)", "Revenue / Paying Users", "$7.50", "$7.88", "$8.27", "$5-15"),
    ("Conversion Rate", "Paying / Total Users", "4.0%", "4.5%", "5.0%", "2-7%"),
    ("Day 1 Retention", "Users returning Day 1", "40%", "45%", "48%", "35-50%"),
    ("Day 7 Retention", "Users returning Day 7", "18%", "22%", "25%", "15-25%"),
    ("Day 30 Retention", "Users returning Day 30", "8%", "10%", "12%", "5-12%"),
    ("Paying User Churn", "Lost paying users / total paying", "10%", "9%", "8%", "5-12%"),
    ("Revenue Per Session", "Revenue / Total Sessions", "$0.03", "$0.05", "$0.07", "$0.02-0.10"),
    ("Sessions Per User Per Day", "Total Sessions / DAU", "2.5", "2.8", "3.0", "2-4"),
    ("Gross Margin", "(Rev - COGS) / Revenue", "72%", "75%", "78%", "70-85%"),
    ("Payback Period (months)", "CAC / Monthly ARPU", "7.1", "4.0", "2.4", "< 12"),
    ("Break-even Month", "Cumulative P/L turns positive", "Month 8-10", "—", "—", "6-18 months"),
]

for i, (metric, formula, y1, y2, y3, bench) in enumerate(kpis):
    row = r + 1 + i
    data = [metric, formula, y1, y2, y3, bench]
    write_row(ws_kpi, row, data, alt=i % 2 == 0)
    ws_kpi.cell(row=row, column=1).font = Font(bold=True)
    ws_kpi.cell(row=row, column=1).alignment = Alignment(horizontal="left")
    ws_kpi.cell(row=row, column=2).alignment = Alignment(horizontal="left")
    ws_kpi.cell(row=row, column=6).font = Font(italic=True, color="666666")

# Sensitivity Analysis
sa_r = r + len(kpis) + 3
ws_kpi.merge_cells(f"A{sa_r}:F{sa_r}")
ws_kpi[f"A{sa_r}"].value = "SENSITIVITY ANALYSIS — MONTHLY NET REVENUE AT MONTH 12"
style_header_row(ws_kpi, sa_r, 6, fill=section_fill, font=section_font)

sa_r += 1
sa_headers = ["Conversion \\ ARPPU", "$5.00", "$7.50", "$10.00", "$12.50", "$15.00"]
for i, h in enumerate(sa_headers):
    cell = ws_kpi.cell(row=sa_r, column=i + 1, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

# Sensitivity: assume ~4000 active users at month 12
active_m12 = 4000
arppus = [5.00, 7.50, 10.00, 12.50, 15.00]
conversions = [0.02, 0.03, 0.04, 0.05, 0.07]

for i, conv in enumerate(conversions):
    row = sa_r + 1 + i
    ws_kpi.cell(row=row, column=1, value=f"{conv:.0%} conversion").font = Font(bold=True)
    ws_kpi.cell(row=row, column=1).border = thin_border
    for j, a in enumerate(arppus):
        paying = active_m12 * conv
        net = paying * a * 0.85  # after 15% fee
        cell = ws_kpi.cell(row=row, column=j + 2, value=round(net, 0))
        cell.number_format = usd_whole_fmt
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        # Color code
        if net > 2000:
            cell.fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
        elif net > 1000:
            cell.fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
        else:
            cell.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")

# Column widths
ws_kpi.column_dimensions["A"].width = 32
ws_kpi.column_dimensions["B"].width = 28
for col in range(3, 7):
    ws_kpi.column_dimensions[get_column_letter(col)].width = 16


# ═══════════════════════════════════════════════════════════════
# SHEET 6: SCENARIO PLANNER
# ═══════════════════════════════════════════════════════════════
ws_scenario = wb.create_sheet("Scenario Planner")
ws_scenario.sheet_properties.tabColor = "E67E22"

ws_scenario.merge_cells("A1:G1")
ws_scenario["A1"].value = "SCENARIO PLANNER — CONSERVATIVE / BASE / OPTIMISTIC"
ws_scenario["A1"].font = Font(name="Calibri", bold=True, color=WHITE, size=16)
ws_scenario["A1"].fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
ws_scenario["A1"].alignment = Alignment(horizontal="center")

r = 3
sc_headers = ["Metric", "Conservative", "Base Case", "Optimistic", "Notes"]
style_header_row(ws_scenario, r, 5, fill=header_fill, font=header_font)
for i, h in enumerate(sc_headers):
    ws_scenario.cell(row=r, column=i + 1, value=h)

scenarios = [
    ("Launch Installs", "1,000", "2,000", "5,000", "Depends on launch marketing"),
    ("Monthly Growth Rate", "8%", "12%", "20%", "Organic + paid growth"),
    ("Conversion Rate", "2%", "4%", "7%", "Industry range for casual games"),
    ("ARPPU", "$5.00", "$7.50", "$12.00", "Based on pricing matrix"),
    ("Paying Churn", "15%", "10%", "6%", "Lower = better retention"),
    ("Free Churn", "20%", "15%", "10%", "Content keeps users engaged"),
    ("App Store Fee", "30%", "15%", "15%", "Small Business Program"),
    ("Month 6 Active Users", "1,200", "3,500", "10,000", "Growth dependent"),
    ("Month 12 Active Users", "2,500", "6,000", "25,000", "Viral potential"),
    ("Month 12 Monthly Revenue", "$250", "$2,550", "$15,750", "Net after fees"),
    ("Month 24 Monthly Revenue", "$600", "$8,000", "$45,000", "At scale"),
    ("Year 1 Total Revenue", "$1,500", "$12,000", "$55,000", "Cumulative net"),
    ("Year 2 Total Revenue", "$5,000", "$65,000", "$350,000", "Growth phase"),
    ("Year 3 Total Revenue", "$10,000", "$130,000", "$600,000", "Maturity phase"),
    ("5-Year Total Revenue", "$30,000", "$450,000", "$2,000,000", "Full projection"),
    ("Break-even Month", "Month 18+", "Month 8-10", "Month 3-4", "When cumulative > 0"),
    ("5-Year ROI", "50%", "500%+", "2000%+", "Return on investment"),
]

for i, (metric, cons, base, opt, notes) in enumerate(scenarios):
    row = r + 1 + i
    data = [metric, cons, base, opt, notes]
    write_row(ws_scenario, row, data, alt=i % 2 == 0)
    ws_scenario.cell(row=row, column=1).font = Font(bold=True)
    ws_scenario.cell(row=row, column=1).alignment = Alignment(horizontal="left")

    # Color code scenarios
    ws_scenario.cell(row=row, column=2).font = Font(color=RED)
    ws_scenario.cell(row=row, column=4).font = Font(color=GREEN)
    ws_scenario.cell(row=row, column=5).font = Font(italic=True, color="888888")

# Column widths
ws_scenario.column_dimensions["A"].width = 28
for col in range(2, 6):
    ws_scenario.column_dimensions[get_column_letter(col)].width = 22


# ═══════════════════════════════════════════════════════════════
# SHEET 7: MONTHLY CHECKLIST & MILESTONES
# ═══════════════════════════════════════════════════════════════
ws_check = wb.create_sheet("Milestones & Checklist")
ws_check.sheet_properties.tabColor = "1ABC9C"

ws_check.merge_cells("A1:E1")
ws_check["A1"].value = "MILESTONES & MANAGEMENT CHECKLIST"
ws_check["A1"].font = Font(name="Calibri", bold=True, color=WHITE, size=16)
ws_check["A1"].fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
ws_check["A1"].alignment = Alignment(horizontal="center")

# Revenue Milestones
r = 3
ws_check.merge_cells(f"A{r}:E{r}")
ws_check[f"A{r}"].value = "REVENUE MILESTONES"
style_header_row(ws_check, r, 5, fill=section_fill, font=section_font)

r += 1
m_headers = ["Milestone", "Target", "Target Month", "Status", "Notes"]
for i, h in enumerate(m_headers):
    cell = ws_check.cell(row=r, column=i + 1, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

milestones = [
    ("First Paying User", "1 user", "Month 1", "", "Validate pricing"),
    ("100 Total Users", "100", "Month 2", "", "Early traction"),
    ("$100 Monthly Revenue", "$100/mo", "Month 3-4", "", "Proof of concept"),
    ("1,000 Total Users", "1,000", "Month 4-6", "", "Growth validation"),
    ("$1,000 Monthly Revenue", "$1K/mo", "Month 6-8", "", "Sustainability signal"),
    ("Apple Small Business Program", "< $1M annual", "Month 1", "", "Apply immediately — 15% fee"),
    ("Break-even Month", "Cum P/L > 0", "Month 8-10", "", "Critical milestone"),
    ("$5,000 Monthly Revenue", "$5K/mo", "Month 12-18", "", "Part-time income level"),
    ("10,000 Active Users", "10K MAU", "Month 12-18", "", "Community scale"),
    ("$10,000 Monthly Revenue", "$10K/mo", "Month 18-24", "", "Full-time viable"),
    ("$1M Cumulative Revenue", "$1M total", "Month 36-48", "", "Major milestone"),
]

for i, (ms, target, tgt_month, status, notes) in enumerate(milestones):
    row = r + 1 + i
    data = [ms, target, tgt_month, status, notes]
    write_row(ws_check, row, data, alt=i % 2 == 0)
    ws_check.cell(row=row, column=1).font = Font(bold=True)
    ws_check.cell(row=row, column=1).alignment = Alignment(horizontal="left")

# Monthly checklist
cr = r + len(milestones) + 3
ws_check.merge_cells(f"A{cr}:E{cr}")
ws_check[f"A{cr}"].value = "MONTHLY MANAGEMENT CHECKLIST"
style_header_row(ws_check, cr, 5, fill=section_fill, font=section_font)

cr += 1
checklist = [
    ("Review App Store Analytics", "Weekly", "Downloads, ratings, reviews, crash reports"),
    ("Check Revenue Dashboard", "Daily", "IAP revenue, refund rate, chargebacks"),
    ("Monitor Conversion Funnel", "Weekly", "Install → Register → First Purchase → Repeat"),
    ("Update User Cohort Analysis", "Monthly", "Retention curves by install week"),
    ("Review Churn Reasons", "Monthly", "Survey lapsed users, analyze drop-off points"),
    ("A/B Test Pricing", "Quarterly", "Test different price points and bundles"),
    ("Update Financial Model", "Monthly", "Actuals vs. forecast, adjust projections"),
    ("Review Season Pass Performance", "Per Season", "Completion rates, upgrade rates"),
    ("Competitor Analysis", "Quarterly", "New entrants, pricing changes, features"),
    ("Tax & Compliance Review", "Quarterly", "Sales tax, VAT, app store settlements"),
    ("Marketing ROI Analysis", "Monthly", "CAC by channel, ROAS, attribution"),
    ("User Feedback Review", "Weekly", "App store reviews, support tickets, social media"),
    ("Feature Prioritization", "Monthly", "Revenue impact vs. development cost"),
    ("Cash Flow Forecast Update", "Monthly", "60-90 day cash runway, payment timing"),
]

c_headers = ["Task", "Frequency", "Details", "Done?", "Notes"]
for i, h in enumerate(c_headers):
    cell = ws_check.cell(row=cr, column=i + 1, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

for i, (task, freq, details) in enumerate(checklist):
    row = cr + 1 + i
    data = [task, freq, details, "", ""]
    write_row(ws_check, row, data, alt=i % 2 == 0)
    ws_check.cell(row=row, column=1).alignment = Alignment(horizontal="left")
    ws_check.cell(row=row, column=3).alignment = Alignment(horizontal="left")

# Column widths
ws_check.column_dimensions["A"].width = 32
ws_check.column_dimensions["B"].width = 16
ws_check.column_dimensions["C"].width = 45
ws_check.column_dimensions["D"].width = 12
ws_check.column_dimensions["E"].width = 25


# ═══════════════════════════════════════════════════════════════
# SHEET 8: APP STORE FEE OPTIMIZER
# ═══════════════════════════════════════════════════════════════
ws_fees = wb.create_sheet("Fee Optimizer")
ws_fees.sheet_properties.tabColor = "3498DB"

ws_fees.merge_cells("A1:F1")
ws_fees["A1"].value = "APP STORE FEE OPTIMIZATION GUIDE"
ws_fees["A1"].font = Font(name="Calibri", bold=True, color=WHITE, size=16)
ws_fees["A1"].fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
ws_fees["A1"].alignment = Alignment(horizontal="center")

r = 3
ws_fees.merge_cells(f"A{r}:F{r}")
ws_fees[f"A{r}"].value = "APPLE APP STORE SMALL BUSINESS PROGRAM"
style_header_row(ws_fees, r, 6, fill=section_fill, font=section_font)

r += 1
apple_info = [
    ("Eligibility", "Earn < $1M in annual proceeds across all apps"),
    ("Commission Rate", "15% (vs. standard 30%)"),
    ("Savings", "50% reduction in fees"),
    ("Application", "Apply at developer.apple.com/enroll"),
    ("Renewal", "Auto-renews if you stay under $1M"),
    ("If You Exceed $1M", "Standard 30% rate applies for rest of year"),
    ("Subscriptions", "15% after Year 1 subscriber (regardless of program)"),
]

for i, (label, detail) in enumerate(apple_info):
    row = r + i
    ws_fees.cell(row=row, column=1, value=label).font = Font(bold=True)
    ws_fees.cell(row=row, column=1).border = thin_border
    ws_fees.merge_cells(f"B{row}:F{row}")
    ws_fees.cell(row=row, column=2, value=detail).border = thin_border
    if i % 2 == 0:
        for c in range(1, 7):
            ws_fees.cell(row=row, column=c).fill = alt_fill

r2 = r + len(apple_info) + 2
ws_fees.merge_cells(f"A{r2}:F{r2}")
ws_fees[f"A{r2}"].value = "GOOGLE PLAY FEE STRUCTURE"
style_header_row(ws_fees, r2, 6, fill=section_fill, font=section_font)

r2 += 1
google_info = [
    ("Service Fee Tier 1", "15% on first $1M of annual revenue"),
    ("Service Fee Tier 2", "30% on revenue above $1M"),
    ("Subscriptions", "15% for all subscription revenue (as of 2022)"),
    ("Key Advantage", "Google's 15% tier applies to ALL developers, no application needed"),
    ("Media Content", "10% for ebooks, music streaming"),
    ("Payment", "Net-30 payment terms via Google payments"),
]

for i, (label, detail) in enumerate(google_info):
    row = r2 + i
    ws_fees.cell(row=row, column=1, value=label).font = Font(bold=True)
    ws_fees.cell(row=row, column=1).border = thin_border
    ws_fees.merge_cells(f"B{row}:F{row}")
    ws_fees.cell(row=row, column=2, value=detail).border = thin_border
    if i % 2 == 0:
        for c in range(1, 7):
            ws_fees.cell(row=row, column=c).fill = alt_fill

# Fee savings calculator
r3 = r2 + len(google_info) + 2
ws_fees.merge_cells(f"A{r3}:F{r3}")
ws_fees[f"A{r3}"].value = "FEE SAVINGS CALCULATOR — 15% vs 30%"
style_header_row(ws_fees, r3, 6, fill=section_fill, font=section_font)

r3 += 1
calc_headers = ["Annual Gross Revenue", "At 30% Fee", "At 15% Fee", "You Keep (30%)", "You Keep (15%)", "Annual Savings"]
for i, h in enumerate(calc_headers):
    cell = ws_fees.cell(row=r3, column=i + 1, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

revenues = [5000, 10000, 25000, 50000, 100000, 250000, 500000, 750000, 1000000]
for i, rev in enumerate(revenues):
    row = r3 + 1 + i
    fee_30 = rev * 0.30
    fee_15 = rev * 0.15
    keep_30 = rev - fee_30
    keep_15 = rev - fee_15
    savings = keep_15 - keep_30
    data = [rev, fee_30, fee_15, keep_30, keep_15, savings]
    fmts = [usd_whole_fmt] * 6
    write_row(ws_fees, row, data, fmt=fmts, alt=i % 2 == 0)
    # Highlight savings
    ws_fees.cell(row=row, column=6).fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    ws_fees.cell(row=row, column=6).font = Font(bold=True, color=GREEN)

# Column widths
ws_fees.column_dimensions["A"].width = 24
for col in range(2, 7):
    ws_fees.column_dimensions[get_column_letter(col)].width = 20


# ═══════════════════════════════════════════════════════════════
# CHARTS on a dedicated sheet
# ═══════════════════════════════════════════════════════════════
ws_charts = wb.create_sheet("Charts")
ws_charts.sheet_properties.tabColor = GOLD

# We need data references from the forecast sheet
# The forecast data starts at row fr+1 in the "60-Month Forecast" sheet
# Columns: A=Month, B=New Installs, D=Free Users, E=Paying Users, F=Total Active,
#           H=Gross Rev, J=Net Rev, O=Net P/L, P=Cum P/L

forecast_sheet_title = "'60-Month Forecast'"

# Chart 1: User Growth (Free vs Paying)
chart1 = LineChart()
chart1.title = "User Growth — Free vs Paying Users"
chart1.style = 10
chart1.y_axis.title = "Users"
chart1.x_axis.title = "Month"
chart1.width = 30
chart1.height = 15

# We can't reference another sheet easily with openpyxl charts on a separate sheet,
# so let's put charts on the forecast sheet itself instead
# Actually, let's create a small data mirror on the Charts sheet

# Copy key data to Charts sheet for chart references
charts_data_start = 2
ws_charts.cell(row=1, column=1, value="Month")
ws_charts.cell(row=1, column=2, value="Free Users")
ws_charts.cell(row=1, column=3, value="Paying Users")
ws_charts.cell(row=1, column=4, value="Total Active")
ws_charts.cell(row=1, column=5, value="Gross Revenue")
ws_charts.cell(row=1, column=6, value="Net Revenue")
ws_charts.cell(row=1, column=7, value="Net Profit/Loss")
ws_charts.cell(row=1, column=8, value="Cumulative P/L")
ws_charts.cell(row=1, column=9, value="New Installs")

for m in range(1, 61):
    r_src = fr + m  # row in forecast sheet
    r_dst = charts_data_start + m - 1
    ws_charts.cell(row=r_dst, column=1, value=m)
    # Reference formulas from forecast sheet
    ws_charts.cell(row=r_dst, column=2).value = f"='{ws_forecast.title}'!D{r_src}"
    ws_charts.cell(row=r_dst, column=3).value = f"='{ws_forecast.title}'!E{r_src}"
    ws_charts.cell(row=r_dst, column=4).value = f"='{ws_forecast.title}'!F{r_src}"
    ws_charts.cell(row=r_dst, column=5).value = f"='{ws_forecast.title}'!H{r_src}"
    ws_charts.cell(row=r_dst, column=6).value = f"='{ws_forecast.title}'!J{r_src}"
    ws_charts.cell(row=r_dst, column=7).value = f"='{ws_forecast.title}'!O{r_src}"
    ws_charts.cell(row=r_dst, column=8).value = f"='{ws_forecast.title}'!P{r_src}"
    ws_charts.cell(row=r_dst, column=9).value = f"='{ws_forecast.title}'!B{r_src}"

data_end = charts_data_start + 59

# Chart 1: User Growth
chart1 = LineChart()
chart1.title = "User Growth Over 60 Months"
chart1.style = 10
chart1.y_axis.title = "Users"
chart1.x_axis.title = "Month"
chart1.width = 32
chart1.height = 16

cats = Reference(ws_charts, min_col=1, min_row=charts_data_start, max_row=data_end)
for col, name, color in [(2, "Free Users", BLUE), (3, "Paying Users", GREEN), (4, "Total Active", GOLD)]:
    vals = Reference(ws_charts, min_col=col, min_row=1, max_row=data_end)
    chart1.add_data(vals, titles_from_data=True)
    chart1.set_categories(cats)

chart1.series[0].graphicalProperties.line.solidFill = BLUE
chart1.series[1].graphicalProperties.line.solidFill = GREEN
chart1.series[2].graphicalProperties.line.solidFill = GOLD

ws_charts.add_chart(chart1, "K2")

# Chart 2: Revenue
chart2 = LineChart()
chart2.title = "Monthly Revenue (Gross vs Net vs Profit)"
chart2.style = 10
chart2.y_axis.title = "USD ($)"
chart2.x_axis.title = "Month"
chart2.width = 32
chart2.height = 16

cats2 = Reference(ws_charts, min_col=1, min_row=charts_data_start, max_row=data_end)
for col in [5, 6, 7]:
    vals = Reference(ws_charts, min_col=col, min_row=1, max_row=data_end)
    chart2.add_data(vals, titles_from_data=True)
chart2.set_categories(cats2)

chart2.series[0].graphicalProperties.line.solidFill = GOLD
chart2.series[1].graphicalProperties.line.solidFill = GREEN
chart2.series[2].graphicalProperties.line.solidFill = RED

ws_charts.add_chart(chart2, "K20")

# Chart 3: Cumulative P/L
chart3 = LineChart()
chart3.title = "Cumulative Profit/Loss Over 60 Months"
chart3.style = 10
chart3.y_axis.title = "USD ($)"
chart3.x_axis.title = "Month"
chart3.width = 32
chart3.height = 16

cats3 = Reference(ws_charts, min_col=1, min_row=charts_data_start, max_row=data_end)
vals3 = Reference(ws_charts, min_col=8, min_row=1, max_row=data_end)
chart3.add_data(vals3, titles_from_data=True)
chart3.set_categories(cats3)
chart3.series[0].graphicalProperties.line.solidFill = GREEN

ws_charts.add_chart(chart3, "K38")

# Hide the raw data columns
for col in range(1, 10):
    ws_charts.column_dimensions[get_column_letter(col)].width = 0.5


# ═══════════════════════════════════════════════════════════════
# FINAL: Save
# ═══════════════════════════════════════════════════════════════
output_path = "e:/Donkey.Marble.Racing/Donkey_Marble_Racing_Financial_Forecast.xlsx"
wb.save(output_path)
print(f"Workbook saved to: {output_path}")
print(f"Sheets: {wb.sheetnames}")
print("Done!")
